"""Export service - builds XLSX (styled) and CSV exports per agent / all agents.

Structure per sheet (xelplus-style clean design, matching scripts/generate_demo_xlsx.py):
- Title band (dark fill, white bold text)
- Subtitle row (grey)
- Section header bands (colored fill per section)
- Table header rows (bold, colored)
- Data rows with thin borders, alternating banding
- TOTAL / NET rows (bold, light highlight)
- Spacing rows between sections

CSV mirrors samples/sample_Agent*.csv: `=== TABLE: NAME ===` separators with
plain comma-separated rows.  agent=all combines every agent into one stream.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
import csv
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from db.models import (
    AuditLog,
    BankTransaction,
    Budget,
    CashFlowProjection,
    CashPosition,
    ChartOfAccount,
    ChequeRegistry,
    ComplianceDeadline,
    Contact,
    DepreciationSchedule,
    EobiRate,
    ExchangeRate,
    FixedAsset,
    FlaggedEntry,
    JournalEntry,
    LoanPaymentSchedule,
    PettyCashFund,
    PettyCashTransaction,
    PrepaidExpense,
    ReceiptExtraction,
    ReconciliationMatch,
    ReconciliationRun,
    RetainedEarnings,
    StatutoryRegister,
    SystemBackupLog,
    SystemConfig,
    TaxFiling,
    TaxRate,
    UserRole,
)

# ---------------------------------------------------------------------------
# palette (CA professional, not garish) - copied from generate_demo_xlsx.py
# ---------------------------------------------------------------------------
NAVY = "1F3864"   # title band
BLUE = "2E75B6"   # section header
LIGHT = "DEEBF7"  # alt row band
TOTAL = "FFF2CC"  # total highlight
WHITE = "FFFFFF"
GREY = "808080"
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

F_TITLE = Font(bold=True, size=14, color=WHITE)
F_SUB = Font(italic=True, size=10, color=GREY)
F_SECTION = Font(bold=True, size=11, color=WHITE)
F_HEAD = Font(bold=True, size=10, color=NAVY)
F_DATA = Font(size=10)
F_TOTAL = Font(bold=True, size=10, color=NAVY)
F_MONO = Font(size=10, name="Consolas")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_SECT = PatternFill("solid", fgColor=BLUE)
FILL_ALT = PatternFill("solid", fgColor=LIGHT)
FILL_TOTAL = PatternFill("solid", fgColor=TOTAL)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

# Merge span for title / section bands (max column count across exports)
SPAN = 12


# ---------------------------------------------------------------------------
# styling helpers (copied from scripts/generate_demo_xlsx.py)
# ---------------------------------------------------------------------------

def title_band(ws, row, text, span=SPAN):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE
    c.fill = FILL_NAVY
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24


def sub_row(ws, row, text, span=SPAN):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SUB
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16


def section_band(ws, row, text, span=SPAN):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    c.fill = FILL_SECT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_ALT
        c.border = BORDER
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 18


def data_row(ws, row, values, alt=False, mono_cols=()):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_MONO if i in mono_cols else F_DATA
        c.border = BORDER
        c.alignment = LEFT if i == 1 else RIGHT if isinstance(v, (int, float, Decimal)) else LEFT
        if alt:
            c.fill = FILL_ALT


def total_row(ws, row, values, span=SPAN):
    for i in range(1, span + 1):
        ws.cell(row=row, column=i).fill = FILL_TOTAL
        ws.cell(row=row, column=i).border = BORDER
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_TOTAL
        c.alignment = CENTER if i == 1 else RIGHT


# ---------------------------------------------------------------------------
# registry
# ---------------------------------------------------------------------------

AGENT_ORDER = [
    "daily-entry",
    "ledger",
    "reconciliation",
    "month-end",
    "year-end",
    "cost-budgeting",
    "tax",
    "audit",
    "advisory",
    "system-admin",
]

AGENT_META = {
    "daily-entry": {"title": "DAILY ENTRY", "number": 1, "sheet": "Daily Entry"},
    "ledger": {"title": "LEDGER & MASTER DATA", "number": 2, "sheet": "Ledger & Master Data"},
    "reconciliation": {"title": "RECONCILIATION & BANKING", "number": 3, "sheet": "Reconciliation & Banking"},
    "month-end": {"title": "MONTH-END REPORTING", "number": 4, "sheet": "Month-End Reporting"},
    "year-end": {"title": "YEAR-END & FINANCIAL STATEMENTS", "number": 5, "sheet": "Year-End & Financials"},
    "cost-budgeting": {"title": "COST, ADVANCED ACCOUNTING & BUDGETING", "number": 6, "sheet": "Cost & Budgeting"},
    "tax": {"title": "TAX", "number": 7, "sheet": "Tax"},
    "audit": {"title": "AUDIT & REGULATORY", "number": 8, "sheet": "Audit & Regulatory"},
    "advisory": {"title": "ADVISORY", "number": 9, "sheet": "Advisory"},
    "system-admin": {"title": "SYSTEM ADMINISTRATION", "number": 10, "sheet": "System Admin"},
}


# ---------------------------------------------------------------------------
# row helpers
# ---------------------------------------------------------------------------

def _row_to_dict(obj) -> dict:
    """Convert a SQLAlchemy model instance to a plain dict (all columns)."""
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}


def _all_rows(db: Session, model) -> list[dict]:
    return [_row_to_dict(r) for r in db.query(model).all()]


def _split_account(value) -> tuple[str, str]:
    """Split '1000-Cash' into (code, name). Returns (value, value) if no hyphen."""
    value = str(value)
    parts = value.split("-", 1)
    if len(parts) == 2 and parts[0].strip().isdigit():
        return parts[0].strip(), parts[1].strip()
    return value, value


def _code_prefix(account) -> str:
    code, _ = _split_account(account)
    return code[:1] if code else ""


def _is_revenue(account) -> bool:
    name = str(account).lower()
    if any(k in name for k in ("revenue", "income", "sales revenue")):
        return True
    return _code_prefix(account) == "4"


def _is_expense(account) -> bool:
    name = str(account).lower()
    if any(k in name for k in ("expense", "cost of goods", "cogs", "cost of sales")):
        return True
    return _code_prefix(account) in ("5", "6", "8")


def _is_cash(account) -> bool:
    name = str(account).lower()
    return any(k in name for k in ("cash", "bank"))


def _is_liability_or_equity(account) -> bool:
    return _code_prefix(account) in ("2", "3")


# ---------------------------------------------------------------------------
# query functions (each returns list of dict rows)
# ---------------------------------------------------------------------------

def _cash_position(db: Session) -> list[dict]:
    return _all_rows(db, CashPosition)


def _bank_transactions(db: Session) -> list[dict]:
    return _all_rows(db, BankTransaction)


def _bank_transactions_unmatched(db: Session) -> list[dict]:
    """Bank transactions not present in reconciliation_matches."""
    matched = {r[0] for r in db.query(ReconciliationMatch.bank_txn_id).all()}
    return [
        r for r in _all_rows(db, BankTransaction)
        if r["transaction_id"] not in matched
    ]


def _petty_cash_funds(db: Session) -> list[dict]:
    return _all_rows(db, PettyCashFund)


def _petty_cash_transactions(db: Session) -> list[dict]:
    return _all_rows(db, PettyCashTransaction)


def _receipt_extractions(db: Session) -> list[dict]:
    return _all_rows(db, ReceiptExtraction)


def _chart_of_accounts(db: Session) -> list[dict]:
    return _all_rows(db, ChartOfAccount)


def _journal_entries(db: Session) -> list[dict]:
    return _all_rows(db, JournalEntry)


def _contacts(db: Session) -> list[dict]:
    return _all_rows(db, Contact)


def _fixed_assets(db: Session) -> list[dict]:
    return _all_rows(db, FixedAsset)


def _reconciliation_runs(db: Session) -> list[dict]:
    return _all_rows(db, ReconciliationRun)


def _reconciliation_matches(db: Session) -> list[dict]:
    return _all_rows(db, ReconciliationMatch)


def _cheque_registry(db: Session) -> list[dict]:
    return _all_rows(db, ChequeRegistry)


def _budgets(db: Session) -> list[dict]:
    return _all_rows(db, Budget)


def _loan_payment_schedule(db: Session) -> list[dict]:
    return _all_rows(db, LoanPaymentSchedule)


def _depreciation_schedule(db: Session) -> list[dict]:
    return _all_rows(db, DepreciationSchedule)


def _prepaid_expenses(db: Session) -> list[dict]:
    return _all_rows(db, PrepaidExpense)


def _retained_earnings(db: Session) -> list[dict]:
    return _all_rows(db, RetainedEarnings)


def _cash_flow_projections(db: Session) -> list[dict]:
    return _all_rows(db, CashFlowProjection)


def _exchange_rates(db: Session) -> list[dict]:
    return _all_rows(db, ExchangeRate)


def _tax_rates(db: Session) -> list[dict]:
    return _all_rows(db, TaxRate)


def _tax_filings(db: Session) -> list[dict]:
    """Persisted sales/income tax filings (newest first)."""
    return [
        _row_to_dict(r)
        for r in db.query(TaxFiling).order_by(
            TaxFiling.created_at.desc(), TaxFiling.fiscal_year.desc()
        ).all()
    ]


def _eobi_rates(db: Session) -> list[dict]:
    return _all_rows(db, EobiRate)


def _compliance_deadlines(db: Session) -> list[dict]:
    return _all_rows(db, ComplianceDeadline)


def _audit_log(db: Session) -> list[dict]:
    return _all_rows(db, AuditLog)


def _flagged_entries(db: Session) -> list[dict]:
    return _all_rows(db, FlaggedEntry)


def _statutory_registers(db: Session) -> list[dict]:
    return _all_rows(db, StatutoryRegister)


def _system_config(db: Session) -> list[dict]:
    return _all_rows(db, SystemConfig)


def _user_roles(db: Session) -> list[dict]:
    return _all_rows(db, UserRole)


def _system_backup_log(db: Session) -> list[dict]:
    return _all_rows(db, SystemBackupLog)


# ---------------------------------------------------------------------------
# computed sections (financial statements from journal_entries)
# ---------------------------------------------------------------------------

def _trial_balance(db: Session) -> list[dict]:
    entries = db.query(JournalEntry).filter(JournalEntry.status == "posted").all()
    totals: dict[str, dict] = {}
    for e in entries:
        d_code, d_name = _split_account(e.debit_account)
        info = totals.setdefault(d_code, {"account_name": d_name, "total_debits": Decimal("0"), "total_credits": Decimal("0")})
        info["total_debits"] += e.debit_amount
        c_code, c_name = _split_account(e.credit_account)
        info2 = totals.setdefault(c_code, {"account_name": c_name, "total_debits": Decimal("0"), "total_credits": Decimal("0")})
        info2["total_credits"] += e.credit_amount

    type_map = {"1": "asset", "2": "liability", "3": "equity", "4": "revenue", "5": "expense", "6": "expense", "8": "expense"}
    rows = []
    total_debits = total_credits = Decimal("0")
    for code in sorted(totals):
        info = totals[code]
        total_debits += info["total_debits"]
        total_credits += info["total_credits"]
        rows.append({
            "account_code": code,
            "account_name": info["account_name"],
            "account_type": type_map.get(code[:1], ""),
            "total_debits": info["total_debits"],
            "total_credits": info["total_credits"],
            "balance": info["total_debits"] - info["total_credits"],
        })
    balance = "BALANCED" if total_debits == total_credits else total_debits - total_credits
    rows.append({
        "account_code": "TOTAL",
        "account_name": "",
        "account_type": "",
        "total_debits": total_debits,
        "total_credits": total_credits,
        "balance": balance,
    })
    return rows


def _profit_loss(db: Session) -> list[dict]:
    entries = db.query(JournalEntry).filter(JournalEntry.status == "posted").all()
    revenue_map: dict[str, Decimal] = {}
    expense_map: dict[str, Decimal] = {}
    for e in entries:
        if _is_revenue(e.credit_account):
            revenue_map[e.credit_account] = revenue_map.get(e.credit_account, Decimal("0")) + e.credit_amount
        if _is_expense(e.debit_account):
            expense_map[e.debit_account] = expense_map.get(e.debit_account, Decimal("0")) + e.debit_amount

    rows = [{"account": k, "amount": v, "type": "revenue"} for k, v in sorted(revenue_map.items())]
    rows += [{"account": k, "amount": v, "type": "expense"} for k, v in sorted(expense_map.items())]
    total_revenue = sum(revenue_map.values(), Decimal("0"))
    total_expenses = sum(expense_map.values(), Decimal("0"))
    net = total_revenue - total_expenses
    rows.append({"account": "TOTAL REVENUE", "amount": total_revenue, "type": ""})
    rows.append({"account": "TOTAL EXPENSES", "amount": total_expenses, "type": ""})
    rows.append({"account": "NET INCOME (LOSS)", "amount": net, "type": "profit" if net >= 0 else "loss"})
    return rows


def _balance_sheet(db: Session) -> list[dict]:
    entries = db.query(JournalEntry).filter(JournalEntry.status == "posted").all()
    balances: dict[str, Decimal] = {}
    names: dict[str, str] = {}
    for e in entries:
        d_code, d_name = _split_account(e.debit_account)
        balances[d_code] = balances.get(d_code, Decimal("0")) + e.debit_amount
        names[d_code] = d_name
        c_code, c_name = _split_account(e.credit_account)
        balances[c_code] = balances.get(c_code, Decimal("0")) - e.credit_amount
        names[c_code] = c_name

    rows = []
    total_assets = total_liab = total_equity = Decimal("0")
    for code in sorted(balances):
        net = balances[code]
        amt = abs(net)
        if amt == 0:
            continue
        p = code[:1]
        if p == "1":
            if net >= 0:
                rows.append({"category": "ASSETS", "account": f"{code}-{names.get(code, code)}", "amount": amt})
                total_assets += amt
            else:
                rows.append({"category": "LIABILITIES", "account": f"{code}-{names.get(code, code)} (Overdraft)", "amount": amt})
                total_liab += amt
        elif p == "2":
            if net < 0:
                rows.append({"category": "LIABILITIES", "account": f"{code}-{names.get(code, code)}", "amount": amt})
                total_liab += amt
        elif p == "3":
            if net < 0:
                rows.append({"category": "EQUITY", "account": f"{code}-{names.get(code, code)}", "amount": amt})
                total_equity += amt

    rows.append({"category": "TOTAL ASSETS", "account": "", "amount": total_assets})
    rows.append({"category": "TOTAL LIABILITIES", "account": "", "amount": total_liab})
    rows.append({"category": "TOTAL EQUITY", "account": "", "amount": total_equity})
    rows.append({
        "category": "BALANCED",
        "account": "assets = liabilities + equity",
        "amount": total_assets - (total_liab + total_equity),
    })
    return rows


def _cash_flow(db: Session) -> list[dict]:
    entries = db.query(JournalEntry).filter(JournalEntry.status == "posted").all()
    rows = []
    net_operating = net_investing = net_financing = Decimal("0")
    for e in entries:
        cash_on_debit = _is_cash(e.debit_account)
        cash_on_credit = _is_cash(e.credit_account)
        if cash_on_debit == cash_on_credit:
            continue  # no cash impact
        other = e.credit_account if cash_on_debit else e.debit_account
        amount = e.debit_amount if cash_on_debit else -e.credit_amount
        if _is_revenue(other) or _is_expense(other):
            rows.append({"category": "OPERATING", "description": f"{e.description} ({other})", "amount": amount})
            net_operating += amount
        elif _code_prefix(other) == "1" and not _is_cash(other):
            rows.append({"category": "INVESTING", "description": f"{e.description} ({other})", "amount": amount})
            net_investing += amount
        elif _is_liability_or_equity(other):
            rows.append({"category": "FINANCING", "description": f"{e.description} ({other})", "amount": amount})
            net_financing += amount
        else:
            rows.append({"category": "OPERATING", "description": f"{e.description} ({other})", "amount": amount})
            net_operating += amount

    rows.append({"category": "NET OPERATING", "description": "", "amount": net_operating})
    rows.append({"category": "NET INVESTING", "description": "", "amount": net_investing})
    rows.append({"category": "NET FINANCING", "description": "", "amount": net_financing})
    rows.append({"category": "NET CASH FLOW", "description": "", "amount": net_operating + net_investing + net_financing})
    return rows


def _journal_analysis(db: Session) -> list[dict]:
    """Advisory-style analysis summary computed from journal_entries."""
    entries = db.query(JournalEntry).all()
    posted = [e for e in entries if e.status == "posted"]
    total_debits = sum((e.debit_amount for e in entries), Decimal("0"))
    total_credits = sum((e.credit_amount for e in entries), Decimal("0"))
    revenue = sum(
        (e.credit_amount for e in posted if _is_revenue(e.credit_account)), Decimal("0")
    )
    expenses = sum(
        (e.debit_amount for e in posted if _is_expense(e.debit_account)), Decimal("0")
    )
    return [
        {"metric": "TOTAL JOURNAL ENTRIES", "value": len(entries)},
        {"metric": "TOTAL POSTED ENTRIES", "value": len(posted)},
        {"metric": "TOTAL DEBITS", "value": total_debits},
        {"metric": "TOTAL CREDITS", "value": total_credits},
        {"metric": "IN BALANCE", "value": total_debits == total_credits},
        {"metric": "TOTAL REVENUE", "value": revenue},
        {"metric": "TOTAL EXPENSES", "value": expenses},
        {"metric": "NET INCOME (LOSS)", "value": revenue - expenses},
    ]


# ---------------------------------------------------------------------------
# agent -> sections registry
# ---------------------------------------------------------------------------

AGENT_QUERIES: dict[str, list[tuple[str, Callable[[Session], list[dict]]]]] = {
    "daily-entry": [
        ("CASH POSITION", _cash_position),
        ("BANK TRANSACTIONS", _bank_transactions),
        ("PETTY CASH FUNDS", _petty_cash_funds),
        ("PETTY CASH TRANSACTIONS", _petty_cash_transactions),
        ("RECEIPT EXTRACTIONS", _receipt_extractions),
    ],
    "ledger": [
        ("CHART OF ACCOUNTS", _chart_of_accounts),
        ("JOURNAL ENTRIES", _journal_entries),
        ("CONTACTS", _contacts),
        ("FIXED ASSETS", _fixed_assets),
    ],
    "reconciliation": [
        ("BANK TRANSACTIONS (UNMATCHED)", _bank_transactions_unmatched),
        ("RECONCILIATION RUNS", _reconciliation_runs),
        ("RECONCILIATION MATCHES", _reconciliation_matches),
        ("CHEQUE REGISTRY", _cheque_registry),
    ],
    "month-end": [
        ("BUDGETS", _budgets),
        ("LOAN PAYMENT SCHEDULE", _loan_payment_schedule),
        ("DEPRECIATION SCHEDULE", _depreciation_schedule),
        ("PREPAID EXPENSES", _prepaid_expenses),
    ],
    "year-end": [
        ("TRIAL BALANCE", _trial_balance),
        ("PROFIT & LOSS", _profit_loss),
        ("BALANCE SHEET", _balance_sheet),
        ("CASH FLOW STATEMENT", _cash_flow),
        ("JOURNAL ENTRIES", _journal_entries),
        ("RETAINED EARNINGS", _retained_earnings),
    ],
    "cost-budgeting": [
        ("BUDGETS", _budgets),
        ("CASH FLOW PROJECTIONS", _cash_flow_projections),
        ("EXCHANGE RATES", _exchange_rates),
    ],
    "tax": [
        ("TAX RATES", _tax_rates),
        ("EOBI RATES", _eobi_rates),
        ("TAX FILINGS", _tax_filings),
        ("COMPLIANCE DEADLINES", _compliance_deadlines),
    ],
    "audit": [
        ("AUDIT LOG", _audit_log),
        ("FLAGGED ENTRIES", _flagged_entries),
        ("STATUTORY REGISTERS", _statutory_registers),
    ],
    "advisory": [
        ("JOURNAL ENTRY ANALYSIS", _journal_analysis),
        ("RETAINED EARNINGS", _retained_earnings),
    ],
    "system-admin": [
        ("AUDIT LOG", _audit_log),
        ("SYSTEM CONFIG", _system_config),
        ("USER ROLES", _user_roles),
        ("SYSTEM BACKUP LOG", _system_backup_log),
    ],
}


# ---------------------------------------------------------------------------
# XLSX rendering
# ---------------------------------------------------------------------------

def _resolve_agents(agent_id: Optional[str]) -> list[str]:
    if agent_id is None or str(agent_id).lower() in ("all", ""):
        return AGENT_ORDER
    aid = str(agent_id)
    if aid not in AGENT_META:
        raise ValueError(
            f"Unknown agent id '{aid}'. Valid ids: {', '.join(AGENT_ORDER)}, or 'all'"
        )
    return [aid]


def _sheet_title(meta: dict) -> str:
    name = f"{meta['number']}. {meta.get('sheet', meta['title'])}"
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    return name[:31]


def _col_widths(headers: list[str], rows: list[dict]) -> list[int]:
    widths = []
    for h in headers:
        w = max(10, len(str(h)))
        for row in rows[:500]:
            v = row.get(h, "")
            w = max(w, len(str(v)))
        widths.append(min(w + 2, 40))
    return widths


def _is_total_row(values: list) -> bool:
    first = str(values[0]) if values else ""
    return (
        first.startswith("TOTAL")
        or first.startswith("NET ")
        or first.startswith("BALANCED")
        or first in ("OPERATING", "INVESTING", "FINANCING")
    )


def _mono_cols(headers: list[str]) -> tuple[int, ...]:
    """Column indices (1-based) that render best in a mono font (IDs/references)."""
    return tuple(
        i + 1 for i, h in enumerate(headers) if str(h).lower().endswith("id") or str(h).lower() == "reference"
    )


def _render_section_xlsx(ws, r: int, section_label: str, rows: list[dict]) -> int:
    section_band(ws, r, section_label, span=SPAN)
    r += 1
    if not rows:
        c = ws.cell(row=r, column=1, value="(No data)")
        c.font = F_SUB
        ws.row_dimensions[r].height = 16
        return r + 2
    headers = list(rows[0].keys())
    widths = _col_widths(headers, rows)
    header_row(ws, r, headers, widths)
    r += 1
    mono = _mono_cols(headers)
    for i, row in enumerate(rows):
        values = [row.get(h, "") for h in headers]
        if _is_total_row(values):
            total_row(ws, r, values, span=len(headers))
        else:
            data_row(ws, r, values, alt=(i % 2 == 1), mono_cols=mono)
        r += 1
    return r + 1  # spacing row after section


def build_xlsx(db: Session, agent_id: Optional[str] = None) -> bytes:
    """Build a styled XLSX workbook.

    agent_id None/'all' -> one sheet per agent; otherwise a single sheet
    with that agent's sections.
    """
    wb = Workbook()
    first = True
    for aid in _resolve_agents(agent_id):
        meta = AGENT_META[aid]
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = _sheet_title(meta)
        ws.sheet_view.showGridLines = False
        r = 1
        title_band(ws, r, f"AI ACCOUNTANT  |  AGENT {meta['number']}: {meta['title']}", span=SPAN)
        r += 1
        sub_row(ws, r, f"Generated: {date.today().isoformat()}  |  Currency: PKR  |  All amounts in rupees", span=SPAN)
        r += 2
        for section_label, query_fn in AGENT_QUERIES[aid]:
            r = _render_section_xlsx(ws, r, section_label, query_fn(db))
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# CSV rendering
# ---------------------------------------------------------------------------

def build_csv(db: Session, agent_id: Optional[str] = None) -> bytes:
    """Build a plain CSV export (samples/sample_Agent*.csv style).

    agent_id None/'all' -> single combined CSV with each agent as a
    `[Agent N: TITLE]` block and `=== TABLE: NAME ===` separators.
    """
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["[AI ACCOUNTANT - EXPORT]"])
    w.writerow([f"[Generated: {date.today().isoformat()}]"])
    for aid in _resolve_agents(agent_id):
        meta = AGENT_META[aid]
        w.writerow([f"[Agent {meta['number']}: {meta['title']}]"])
        w.writerow([])
        for section_label, query_fn in AGENT_QUERIES[aid]:
            w.writerow([f"=== TABLE: {section_label} ==="])
            rows = query_fn(db)
            if not rows:
                w.writerow(["(No data)"])
                w.writerow([])
                continue
            headers = list(rows[0].keys())
            w.writerow(headers)
            for row in rows:
                w.writerow([row.get(h, "") for h in headers])
            w.writerow([])
    return out.getvalue().encode("utf-8")
