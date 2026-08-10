"""Generate a professional CA-style XLSX export demo with fake data.

Structure per sheet (xelplus-style clean design):
- Title band (dark fill, white bold text)
- Subtitle row (grey)
- Section header bands (colored fill per agent)
- Table header rows (bold, colored)
- Data rows with thin borders, alternating banding
- TOTAL rows (bold, light highlight)
- Spacing rows between sections
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- palette (CA professional, not garish) ----------
NAVY   = "1F3864"   # title band
BLUE   = "2E75B6"   # section header
LIGHT  = "DEEBF7"   # alt row band
TOTAL  = "FFF2CC"   # total highlight
WHITE  = "FFFFFF"
GREY   = "808080"
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

THIN_LEFT = Border(left=Side(style="thick", color=NAVY))  # CA side rule on account column

F_TITLE  = Font(bold=True, size=14, color=WHITE)
F_SUB    = Font(italic=True, size=10, color=GREY)
F_SECTION= Font(bold=True, size=11, color=WHITE)
F_HEAD   = Font(bold=True, size=10, color=NAVY)
F_DATA   = Font(size=10)
F_TOTAL  = Font(bold=True, size=10, color=NAVY)
F_MONO   = Font(size=10, name="Consolas")

FILL_NAVY  = PatternFill("solid", fgColor=NAVY)
FILL_SECT  = PatternFill("solid", fgColor=BLUE)
FILL_ALT   = PatternFill("solid", fgColor=LIGHT)
FILL_TOTAL = PatternFill("solid", fgColor=TOTAL)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",  vertical="center")
RIGHT  = Alignment(horizontal="right", vertical="center")


def title_band(ws, row, text, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE
    c.fill = FILL_NAVY
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24


def sub_row(ws, row, text, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SUB
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16


def section_band(ws, row, text, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    c.fill = FILL_SECT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_ALT
        c.border = BORDER
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 18


def data_row(ws, row, values, alt=False, mono_cols=()):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_MONO if i in mono_cols else F_DATA
        c.border = BORDER
        c.alignment = LEFT if i == 1 else RIGHT if isinstance(v, (int, float)) else LEFT
        if alt:
            c.fill = FILL_ALT


def total_row(ws, row, values, span=10):
    for i in range(1, span + 1):
        ws.cell(row=row, column=i).fill = FILL_TOTAL
        ws.cell(row=row, column=i).border = BORDER
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_TOTAL
        c.alignment = CENTER if i == 1 else RIGHT


wb = Workbook()

# =====================================================================
# Sheet 1: Daily Entry
# =====================================================================
ws = wb.active
ws.title = "1. Daily Entry"
ws.sheet_view.showGridLines = False

r = 1
title_band(ws, r, "AI ACCOUNTANT  |  AGENT 1: DAILY ENTRY"); r += 1
sub_row(ws, r, "Generated: 2026-07-31  |  Currency: PKR  |  All amounts in rupees"); r += 2

# --- CASH POSITION ---
section_band(ws, r, "CASH POSITION", span=10); r += 1
header_row(ws, r, ["as_of_date", "account_id", "account_name", "opening", "debits", "credits", "closing", "currency"], [12, 18, 26, 14, 14, 14, 14, 10]); r += 1
data_row(ws, r, ["2026-07-31", "1000-Cash", "Cash on Hand", 500000, 150000, 165000, 485000, "PKR"], mono_cols=(1, 2)); r += 1
data_row(ws, r, ["2026-07-31", "1100-Bank Account", "Bank Account - HBL", 2000000, 850000, 720000, 2130000, "PKR"], alt=True, mono_cols=(1, 2)); r += 1
total_row(ws, r, ["TOTAL CASH", "", "", 2500000, 1000000, 885000, 2615000, ""]); r += 2

# --- BANK TRANSACTIONS ---
section_band(ws, r, "BANK TRANSACTIONS", span=10); r += 1
header_row(ws, r, ["txn_id", "date", "description", "amount", "type", "status", "reference", "balance_after"], [18, 12, 30, 14, 10, 12, 14, 16]); r += 1
for i, row in enumerate([
    ["BNK-001", "2026-07-05", "HBL Payment - Rent", 65000, "debit", "cleared", "CHQ-101", 1935000],
    ["BNK-002", "2026-07-10", "HBL Receipt - RetailMart", 250000, "credit", "cleared", "INV-045", 2185000],
    ["BNK-003", "2026-07-15", "Utilities Bill Payment", 12000, "debit", "cleared", "", 2173000],
    ["BNK-004", "2026-07-28", "Bank Charges July", 500, "debit", "cleared", "", 2172500],
], start=0):
    data_row(ws, r, row, alt=(i % 2 == 1), mono_cols=(1, 7)); r += 1
total_row(ws, r, ["TOTAL", "", "", 327500, "", "", "", ""]); r += 2

# --- PETTY CASH ---
section_band(ws, r, "PETTY CASH", span=10); r += 1
header_row(ws, r, ["fund_id", "fund_name", "custodian", "current_balance", "currency", "threshold", "", "", ""], [12, 20, 14, 16, 10, 12, 10, 10, 10]); r += 1
data_row(ws, r, ["PC-001", "Office Petty Cash", "Ahmed Khan", 13000, "PKR", 5000], mono_cols=(1,)); r += 2

# --- RECEIPT EXTRACTIONS ---
section_band(ws, r, "RECEIPT EXTRACTIONS", span=10); r += 1
header_row(ws, r, ["extraction_id", "vendor_name", "total_amount", "date", "currency", "confidence", "needs_approval", "status"], [18, 20, 14, 12, 10, 12, 14, 18]); r += 1
data_row(ws, r, ["REC-20260728-001", "Grocery Store", 5000, "2026-07-28", "PKR", 0.92, "YES", "pending_approval"], mono_cols=(1,)); r += 1
data_row(ws, r, ["REC-20260729-002", "Stationery Mart", 1500, "2026-07-29", "PKR", 0.9, "YES", "pending_approval"], alt=True, mono_cols=(1,)); r += 1

# =====================================================================
# Sheet 2: Ledger & Master Data
# =====================================================================
ws2 = wb.create_sheet("2. Ledger & Master Data")
ws2.sheet_view.showGridLines = False
r = 1
title_band(ws2, r, "AI ACCOUNTANT  |  AGENT 2: LEDGER & MASTER DATA"); r += 1
sub_row(ws2, r, "Chart of Accounts  |  Journal Entries  |  Contacts  |  Fixed Assets"); r += 2

# CHART OF ACCOUNTS
section_band(ws2, r, "CHART OF ACCOUNTS", span=10); r += 1
header_row(ws2, r, ["account_code", "account_name", "account_type", "currency", "is_active"], [18, 26, 14, 10, 10]); r += 1
coa = [
    ["1000-Cash", "Cash on Hand", "asset", "PKR", 1],
    ["1100-Bank Account", "Bank Account - HBL", "asset", "PKR", 1],
    ["1200-Inventory", "Inventory Stock", "asset", "PKR", 1],
    ["1300-Receivables", "Accounts Receivable", "asset", "PKR", 1],
    ["2000-Payables", "Accounts Payable", "liability", "PKR", 1],
    ["3000-Equity", "Owner's Equity", "equity", "PKR", 1],
    ["4000-Revenue", "Sales Revenue", "revenue", "PKR", 1],
    ["5000-COGS", "Cost of Goods Sold", "expense", "PKR", 1],
    ["6000-Rent", "Office Rent", "expense", "PKR", 1],
    ["6100-Salary", "Salaries & Wages", "expense", "PKR", 1],
]
for i, row in enumerate(coa):
    data_row(ws2, r, row, alt=(i % 2 == 1), mono_cols=(1,)); r += 1
r += 1

# JOURNAL ENTRIES
section_band(ws2, r, "JOURNAL ENTRIES", span=10); r += 1
header_row(ws2, r, ["entry_id", "posted_date", "description", "ref", "contact", "debit_acct", "debit", "credit_acct", "credit"], [18, 12, 30, 12, 12, 16, 12, 16, 12]); r += 1
jes = [
    ["JE-202607-001", "2026-07-01", "Office Rent July 2026", "", "", "6000-Rent", 65000, "1100-Bank Account", 65000],
    ["JE-202607-002", "2026-07-10", "Sales Revenue - RetailMart", "INV-045", "CONT-002", "1100-Bank Account", 250000, "4000-Revenue", 250000],
    ["JE-202607-003", "2026-07-12", "Purchase Inventory", "PO-101", "CONT-001", "1200-Inventory", 180000, "2000-Payables", 180000],
    ["JE-202607-004", "2026-07-28", "Salary Disbursement July", "", "", "6100-Salary", 450000, "1100-Bank Account", 450000],
]
for i, row in enumerate(jes):
    data_row(ws2, r, row, alt=(i % 2 == 1), mono_cols=(1, 5)); r += 1
total_row(ws2, r, ["TOTAL", "", "", "", "", "", 945000, "", 945000]); r += 2

# CONTACTS
section_band(ws2, r, "CONTACTS", span=10); r += 1
header_row(ws2, r, ["contact_id", "contact_name", "type", "phone", "email", "tax_id"], [14, 24, 10, 14, 26, 14]); r += 1
for i, row in enumerate([
    ["CONT-001", "TechSolutions Pvt Ltd", "vendor", "021-34920001", "info@techsolutions.pk", "NTN-1234567-1"],
    ["CONT-002", "RetailMart Karachi", "customer", "021-34920002", "orders@retailmart.pk", "NTN-2345678-1"],
    ["CONT-003", "GreenEnergy Corp", "vendor", "042-34920003", "billing@greenenergy.pk", "NTN-3456789-1"],
    ["CONT-004", "Prime Traders", "customer", "051-34920004", "info@primetraders.pk", "NTN-4567890-1"],
]):
    data_row(ws2, r, row, alt=(i % 2 == 1), mono_cols=(1,)); r += 1
r += 1

# FIXED ASSETS
section_band(ws2, r, "FIXED ASSETS", span=10); r += 1
header_row(ws2, r, ["asset_id", "asset_name", "category", "purchase_cost", "purchase_date", "life_yrs", "method", "book_value"], [12, 22, 14, 14, 14, 10, 16, 14]); r += 1
for i, row in enumerate([
    ["FA-001", "Delivery Van", "Vehicle", 2500000, "2025-01-15", 10, "declining_balance", 2000000],
    ["FA-002", "Office Computers", "Computer/IT", 800000, "2025-06-01", 5, "straight_line", 600000],
]):
    data_row(ws2, r, row, alt=(i % 2 == 1), mono_cols=(1,)); r += 1

# =====================================================================
# Sheet 3: Financial Statements
# =====================================================================
ws3 = wb.create_sheet("3. Financial Statements")
ws3.sheet_view.showGridLines = False
r = 1
title_band(ws3, r, "AI ACCOUNTANT  |  AGENT 4: YEAR-END & FINANCIAL STATEMENTS"); r += 1
sub_row(ws3, r, "Trial Balance  |  Profit & Loss  |  Balance Sheet  |  Cash Flow"); r += 2

# TRIAL BALANCE
section_band(ws3, r, "TRIAL BALANCE  (in_balance: TRUE)", span=10); r += 1
header_row(ws3, r, ["account_code", "account_name", "type", "total_debits", "total_credits", "balance"], [14, 20, 10, 14, 14, 14]); r += 1
for i, row in enumerate([
    ["1000", "Cash", "Asset", 150000, 165000, -15000],
    ["1100", "Bank Account", "Asset", 850000, 720000, 130000],
    ["1200", "Inventory", "Asset", 180000, 0, 180000],
    ["2000", "Payables", "Liability", 0, 180000, -180000],
    ["3000", "Equity", "Equity", 0, 500000, -500000],
    ["4000", "Revenue", "Revenue", 0, 250000, -250000],
    ["6000", "Rent", "Expense", 65000, 0, 65000],
    ["6100", "Salary", "Expense", 450000, 0, 450000],
]):
    data_row(ws3, r, row, alt=(i % 2 == 1), mono_cols=(1,)); r += 1
total_row(ws3, r, ["TOTAL", "", "", 1695000, 1695000, "BALANCED"]); r += 2

# PROFIT & LOSS
section_band(ws3, r, "PROFIT & LOSS  (Jan-Jul 2026)", span=10); r += 1
header_row(ws3, r, ["account", "amount", "type", "", "", ""], [24, 14, 14, 10, 10, 10]); r += 1
data_row(ws3, r, ["4000-Revenue", 250000, "revenue"]); r += 1
data_row(ws3, r, ["6000-Rent", 65000, "expense"], alt=True); r += 1
data_row(ws3, r, ["6100-Salary", 450000, "expense"]); r += 1
total_row(ws3, r, ["TOTAL REVENUE", 250000, ""]); r += 1
total_row(ws3, r, ["TOTAL EXPENSES", 515000, ""]); r += 1
total_row(ws3, r, ["NET INCOME (LOSS)", -265000, "LOSS"]); r += 2

# BALANCE SHEET
section_band(ws3, r, "BALANCE SHEET  (balanced: TRUE)", span=10); r += 1
header_row(ws3, r, ["category", "account", "amount", "", "", ""], [14, 24, 14, 10, 10, 10]); r += 1
for i, row in enumerate([
    ["ASSETS", "Cash", 150000],
    ["", "Bank Account", 850000],
    ["", "Inventory", 180000],
    ["", "Accounts Receivable", 250000],
    ["TOTAL ASSETS", "", 1430000],
    ["LIABILITIES", "Accounts Payable", 180000],
    ["TOTAL LIABILITIES", "", 180000],
    ["EQUITY", "Owner's Equity", 500000],
    ["", "Retained Earnings", -265000],
    ["TOTAL EQUITY", "", 235000],
]):
    if row[0].startswith("TOTAL"):
        total_row(ws3, r, row); r += 1
    else:
        data_row(ws3, r, row, alt=(i % 2 == 1)); r += 1
r += 1

# CASH FLOW
section_band(ws3, r, "CASH FLOW STATEMENT", span=10); r += 1
header_row(ws3, r, ["category", "description", "amount", "", "", ""], [14, 30, 14, 10, 10, 10]); r += 1
for i, row in enumerate([
    ["OPERATING", "Revenue Receipt", 250000],
    ["", "Rent Payment", -65000],
    ["", "Salary Payment", -450000],
    ["NET OPERATING", "", -265000],
    ["INVESTING", "", 0],
    ["FINANCING", "Loan Repayment", -100000],
    ["NET CASH FLOW", "", -365000],
]):
    if row[0].startswith("NET") or row[0] in ("INVESTING", "FINANCING"):
        total_row(ws3, r, row); r += 1
    else:
        data_row(ws3, r, row, alt=(i % 2 == 1)); r += 1

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "samples", "AI_Accountant_Demo_Export.xlsx")
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("Saved:", out)
